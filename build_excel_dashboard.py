import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

FONT = "Arial"
NAVY = "1F3864"
ACCENT = "2C5F7C"
LIGHT = "EDF2F7"
WHITE = "FFFFFF"

df = pd.read_csv("data/superstore_clean.csv")
df.columns = [c.strip() for c in df.columns]
df["Order Date"] = pd.to_datetime(df["Order Date"], format="%m/%d/%Y")
df = df.dropna(subset=["Order Date", "Sales", "Order Quantity"])
df = df[df["Order Date"] >= "2009-01-01"]
df["Month"] = df["Order Date"].dt.to_period("M").dt.to_timestamp()

with open("outputs/forecast_results.json") as f:
    forecast = json.load(f)

total_revenue = df["Sales"].sum()
total_units = df["Order Quantity"].sum()
total_orders = df["Order ID"].nunique()
avg_order_value = total_revenue / total_orders

cat_summary = (
    df.groupby("Product Category")
    .agg(Revenue=("Sales", "sum"), Units=("Order Quantity", "sum"), Profit=("Profit", "sum"))
    .reset_index()
    .sort_values("Revenue", ascending=False)
)

region_summary = (
    df.groupby("Region")["Order Quantity"].sum().reset_index().sort_values("Order Quantity", ascending=False)
)

monthly_trend = (
    df.groupby(["Month", "Product Category"])["Order Quantity"].sum().reset_index()
)
monthly_pivot = monthly_trend.pivot(index="Month", columns="Product Category", values="Order Quantity").fillna(0)

top_sub = (
    df.groupby("Product Sub-Category")["Sales"].sum().reset_index()
    .sort_values("Sales", ascending=False).head(10)
)

wb = Workbook()

# ---------- Dashboard sheet ----------
ws = wb.active
ws.title = "Dashboard"
ws.sheet_view.showGridLines = False

ws.merge_cells("B2:M2")
ws["B2"] = "Retail Demand & Inventory Planning Dashboard"
ws["B2"].font = Font(name=FONT, size=20, bold=True, color=WHITE)
ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
for col in range(2, 14):
    ws.cell(row=2, column=col).fill = PatternFill("solid", fgColor=NAVY)
ws.row_dimensions[2].height = 34

ws.merge_cells("B3:M3")
ws["B3"] = "Saloni Bansal  |  Data source: retail transactions, 2009-2012  |  Built in Excel"
ws["B3"].font = Font(name=FONT, size=10, italic=True, color="666666")
ws.row_dimensions[3].height = 18

def kpi_card(col_start, label, value, fmt="number"):
    col_letter = get_column_letter(col_start)
    col_letter2 = get_column_letter(col_start + 1)
    ws.merge_cells(f"{col_letter}5:{col_letter2}5")
    ws.merge_cells(f"{col_letter}6:{col_letter2}7")
    c1 = ws[f"{col_letter}5"]
    c1.value = label
    c1.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c1.fill = PatternFill("solid", fgColor=ACCENT)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c2 = ws[f"{col_letter}6"]
    c2.value = value
    c2.font = Font(name=FONT, size=18, bold=True, color=NAVY)
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.fill = PatternFill("solid", fgColor=LIGHT)
    if fmt == "millions":
        c2.number_format = '$#,##0.0,,"M"'
    elif fmt == "currency":
        c2.number_format = '$#,##0'
    elif fmt == "number":
        c2.number_format = '#,##0'
    return c2

kpi_card(2, "Total Revenue", total_revenue, "millions")
kpi_card(4, "Total Units Sold", total_units, "number")
kpi_card(6, "Total Orders", total_orders, "number")
kpi_card(8, "Avg Order Value", avg_order_value, "currency")
kpi_card(10, "Categories Tracked", len(cat_summary), "number")

ws.row_dimensions[5].height = 20
ws.row_dimensions[6].height = 26
ws.row_dimensions[7].height = 10

# ---------- Category summary table ----------
start_row = 10
ws.cell(row=start_row, column=2, value="Category Performance").font = Font(name=FONT, size=13, bold=True, color=NAVY)
headers = ["Category", "Revenue", "Units Sold", "Profit"]
for j, h in enumerate(headers):
    c = ws.cell(row=start_row + 1, column=2 + j, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center")

for i, row in enumerate(cat_summary.itertuples(), start=start_row + 2):
    ws.cell(row=i, column=2, value=row._1)
    ws.cell(row=i, column=3, value=row.Revenue).number_format = '$#,##0'
    ws.cell(row=i, column=4, value=row.Units).number_format = '#,##0'
    ws.cell(row=i, column=5, value=row.Profit).number_format = '$#,##0'
    for col in range(2, 6):
        ws.cell(row=i, column=col).font = Font(name=FONT, size=10)
        if (i - start_row) % 2 == 0:
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LIGHT)

cat_bar = BarChart()
cat_bar.title = "Revenue by Category"
cat_bar.style = 10
cat_bar.y_axis.title = "Revenue ($)"
data = Reference(ws, min_col=3, min_row=start_row + 1, max_row=start_row + 1 + len(cat_summary))
cats = Reference(ws, min_col=2, min_row=start_row + 2, max_row=start_row + 1 + len(cat_summary))
cat_bar.add_data(data, titles_from_data=True)
cat_bar.set_categories(cats)
cat_bar.width = 11
cat_bar.height = 7
ws.add_chart(cat_bar, "G10")

# ---------- Region chart ----------
region_start = start_row + len(cat_summary) + 4
ws.cell(row=region_start, column=2, value="Units Sold by Region").font = Font(name=FONT, size=13, bold=True, color=NAVY)
ws.cell(row=region_start + 1, column=2, value="Region").font = Font(name=FONT, size=10, bold=True, color=WHITE)
ws.cell(row=region_start + 1, column=2).fill = PatternFill("solid", fgColor=ACCENT)
ws.cell(row=region_start + 1, column=3, value="Units").font = Font(name=FONT, size=10, bold=True, color=WHITE)
ws.cell(row=region_start + 1, column=3).fill = PatternFill("solid", fgColor=ACCENT)
for i, row in enumerate(region_summary.itertuples(), start=region_start + 2):
    ws.cell(row=i, column=2, value=row.Region).font = Font(name=FONT, size=10)
    ws.cell(row=i, column=3, value=row._2).number_format = '#,##0'
    ws.cell(row=i, column=3).font = Font(name=FONT, size=10)

region_bar = BarChart()
region_bar.type = "col"
region_bar.title = "Units Sold by Region"
region_bar.style = 11
rdata = Reference(ws, min_col=3, min_row=region_start + 1, max_row=region_start + 1 + len(region_summary))
rcats = Reference(ws, min_col=2, min_row=region_start + 2, max_row=region_start + 1 + len(region_summary))
region_bar.add_data(rdata, titles_from_data=True)
region_bar.set_categories(rcats)
region_bar.width = 11
region_bar.height = 7
ws.add_chart(region_bar, "G28")

for col, width in zip("BCDEFGHIJKLM", [22,14,14,14,3,14,14,14,14,14,14,14]):
    ws.column_dimensions[col].width = width

ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.print_area = "A1:N45"

# ---------- Monthly Trend sheet ----------
ws2 = wb.create_sheet("Monthly Trend")
ws2.sheet_view.showGridLines = False
ws2["B2"] = "Monthly Demand by Category (Units)"
ws2["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)

ws2.cell(row=4, column=2, value="Month").font = Font(name=FONT, bold=True, color=WHITE)
ws2.cell(row=4, column=2).fill = PatternFill("solid", fgColor=ACCENT)
for j, cat in enumerate(monthly_pivot.columns, start=3):
    c = ws2.cell(row=4, column=j, value=cat)
    c.font = Font(name=FONT, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)

for i, (month, row) in enumerate(monthly_pivot.iterrows(), start=5):
    ws2.cell(row=i, column=2, value=month.strftime("%Y-%m"))
    for j, cat in enumerate(monthly_pivot.columns, start=3):
        ws2.cell(row=i, column=j, value=float(row[cat])).number_format = '#,##0'

trend_chart = LineChart()
trend_chart.title = "Monthly Units Sold by Category"
trend_chart.style = 12
trend_chart.y_axis.title = "Units"
trend_chart.x_axis.title = "Month"
tdata = Reference(ws2, min_col=3, max_col=2 + len(monthly_pivot.columns), min_row=4, max_row=4 + len(monthly_pivot))
tcats = Reference(ws2, min_col=2, min_row=5, max_row=4 + len(monthly_pivot))
trend_chart.add_data(tdata, titles_from_data=True)
trend_chart.set_categories(tcats)
trend_chart.width = 26
trend_chart.height = 12
ws2.add_chart(trend_chart, "B" + str(5 + len(monthly_pivot) + 2))

ws2.column_dimensions["B"].width = 12
for col in "CDE":
    ws2.column_dimensions[col].width = 16

# ---------- Forecast & Reorder sheet ----------
ws3 = wb.create_sheet("Forecast & Reorder")
ws3.sheet_view.showGridLines = False
ws3["B2"] = "Demand Forecast & Reorder Recommendations"
ws3["B2"].font = Font(name=FONT, size=14, bold=True, color=NAVY)
ws3["B3"] = "Forecast method: Holt-Winters exponential smoothing, validated on a 6-month holdout"
ws3["B3"].font = Font(name=FONT, size=10, italic=True, color="666666")

headers3 = ["Category", "Avg Monthly Demand (last 6mo)", "Forecast Accuracy (MAPE)",
            "Fcst Month 1", "Fcst Month 2", "Fcst Month 3", "Recommended Reorder Qty (+15% safety stock)"]
for j, h in enumerate(headers3):
    c = ws3.cell(row=5, column=2 + j, value=h)
    c.font = Font(name=FONT, size=10, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=ACCENT)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
ws3.row_dimensions[5].height = 30

for i, (cat, r) in enumerate(forecast.items(), start=6):
    ws3.cell(row=i, column=2, value=cat)
    ws3.cell(row=i, column=3, value=r["avg_monthly_units_last_6mo"]).number_format = '#,##0'
    ws3.cell(row=i, column=4, value=r["mape_pct"] / 100).number_format = '0.0%'
    for k, val in enumerate(r["next_3mo_forecast"]):
        ws3.cell(row=i, column=5 + k, value=val).number_format = '#,##0'
    ws3.cell(row=i, column=8, value=r["recommended_reorder_qty"]).number_format = '#,##0'
    for col in range(2, 9):
        ws3.cell(row=i, column=col).font = Font(name=FONT, size=10)
        if (i - 5) % 2 == 0:
            ws3.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LIGHT)

for col, width in zip("BCDEFGHI", [18,20,16,12,12,12,26,4]):
    ws3.column_dimensions[col].width = width

wb.save("outputs/retail_dashboard.xlsx")
print("Saved outputs/retail_dashboard.xlsx")
